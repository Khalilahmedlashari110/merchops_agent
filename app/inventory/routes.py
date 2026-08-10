import os
from datetime import datetime
from io import BytesIO
import pandas as pd
from flask import send_file
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

from flask import (
    Blueprint,
    render_template,
    request,
    current_app,
    flash,
    redirect,
    url_for,
    jsonify,
)
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename

from app.inventory.service import (
    CITADEL_DISTRIBUTOR,
    get_default_anchor_date,
    get_citadel_scorecard_data,
    get_brand_inventory_trend,
    apply_requirement_settings,
    apply_requirement_months,
    build_requirement_brand_options,
    get_requirement_settings,
    save_requirement_setting,
    delete_requirement_setting,
    build_brand_summary,
    build_scorecard_analysis,
    get_brand_detail,
    save_presentation_rows_to_db,
    get_presentation_rows_from_db,
    get_presentation_batches,
    delete_presentation_batch,
    build_saved_inventory_brand_tree,
    compare_presentation_batches,
    get_inventory_dashboard_context,
    get_inventory_dashboard_filter_options,
    get_inventory_point_detail,
)
from app.inventory.presentation import (
    read_uploaded_inventory,
    standardize_inventory_dataframe,
    dataframe_to_records,
    get_presentation_summary,
    apply_presentation_filters,
)

inventory_bp = Blueprint("inventory", __name__, url_prefix="/inventory")


def _load_scorecard_context(anchor_str=None, req_months=""):
    anchor_date, rows = _load_scorecard_rows(anchor_str, req_months)
    brand_summary = build_brand_summary(rows)
    analysis = build_scorecard_analysis(rows, brand_summary)
    return anchor_date, rows, brand_summary, analysis


def _load_scorecard_rows(anchor_str=None, req_months=""):
    default_anchor = get_default_anchor_date()
    anchor_str = anchor_str or default_anchor.strftime("%Y-%m-%d")
    anchor_date = datetime.strptime(anchor_str, "%Y-%m-%d").date()
    rows = apply_requirement_settings(get_citadel_scorecard_data(anchor_date), CITADEL_DISTRIBUTOR)
    rows = apply_requirement_months(rows, req_months)
    return anchor_date, rows


def _filter_kpi_export_rows(rows, brand_summary, kpi):
    if kpi == "available":
        return sorted(brand_summary, key=lambda row: row.get("TotalAvailableQty", 0) or 0, reverse=True), "brand"
    if kpi == "in_transit":
        return [row for row in sorted(brand_summary, key=lambda row: row.get("TotalInTransitQty", 0) or 0, reverse=True) if (row.get("TotalInTransitQty", 0) or 0) > 0], "brand"
    if kpi == "req_order":
        return [row for row in sorted(rows, key=lambda row: float(row.get("ReqNewOrderQty", 0) or 0), reverse=True) if float(row.get("ReqNewOrderQty", 0) or 0) > 0], "sku"
    if kpi == "critical":
        return [row for row in sorted(rows, key=lambda row: float(row.get("CoveragePct", 0) or 0)) if float(row.get("CoveragePct", 0) or 0) <= 20 and float(row.get("SKUOldMonths", 0) or 0) >= 10], "sku"
    if kpi == "below_target":
        return [row for row in sorted(rows, key=lambda row: float(row.get("CoveragePct", 0) or 0)) if float(row.get("CoveragePct", 0) or 0) < 100], "sku"
    if kpi == "sales_6m":
        return sorted(rows, key=lambda row: float(row.get("Sales6M", 0) or 0), reverse=True), "sku"
    if kpi == "new_sku":
        return [row for row in sorted(rows, key=lambda row: (str(row.get("dtColLaunch") or ""), str(row.get("Brand") or ""), str(row.get("SKU") or "")), reverse=True) if float(row.get("SKUOldMonths", 0) or 0) < 10], "sku"
    return [], "sku"


def _format_inventory_workbook(writer):
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(color="111827", bold=True, size=14)
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ws in writer.book.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="center")
        for column_cells in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 32)
        if ws.title == "Summary":
            ws["A1"].font = title_font


@inventory_bp.route("/")
@login_required
def index():
    org_id = None if current_user.role_name == "super_admin" else current_user.org_id
    filters = {
        "brand": (request.args.get("brand") or "").strip(),
        "category": (request.args.get("category") or "").strip(),
        "status": (request.args.get("status") or "").strip(),
        "warehouse": (request.args.get("warehouse") or "").strip(),
    }
    snapshot_str = (request.args.get("snapshot") or "").strip() or None
    error = None
    context = {
        "summary": {},
        "products": [],
        "brand_summary": [],
        "category_summary": [],
        "status_summary": [],
        "trend": {"labels": [], "keys": [], "available": [], "value": [], "risk": []},
        "recommendations": [],
        "kpi_details": {},
        "snapshot_date": snapshot_str,
        "filters": filters,
    }
    filter_options = {"brands": [], "categories": [], "statuses": [], "warehouses": []}

    try:
        filter_options = get_inventory_dashboard_filter_options(org_id=org_id)
        context = get_inventory_dashboard_context(
            org_id=org_id,
            snapshot_date=snapshot_str,
            filters=filters,
        )
    except Exception as e:
        error = str(e)

    return render_template(
        "inventory/index.html",
        error=error,
        dashboard=context,
        filter_options=filter_options,
        last_sync=datetime.now().strftime("%d-%b-%Y %I:%M %p"),
    )


@inventory_bp.route("/point-detail")
@login_required
def point_detail():
    org_id = None if current_user.role_name == "super_admin" else current_user.org_id
    filters = {
        "brand": (request.args.get("brand_filter") or "").strip(),
        "category": (request.args.get("category_filter") or "").strip(),
        "status": (request.args.get("status_filter") or "").strip(),
        "warehouse": (request.args.get("warehouse_filter") or "").strip(),
    }
    try:
        detail = get_inventory_point_detail(
            request.args.get("type", ""),
            request.args.get("value", ""),
            org_id=org_id,
            snapshot_date=(request.args.get("snapshot") or "").strip() or None,
            filters=filters,
        )
        return jsonify(detail)
    except Exception as e:
        return jsonify({"success": False, "message": str(e), "rows": []}), 500


@inventory_bp.route("/trend")
@login_required
def trend():
    default_anchor = get_default_anchor_date()
    anchor_str = request.args.get("anchor") or default_anchor.strftime("%Y-%m-%d")
    req_months = (request.args.get("req_months") or "").strip()

    try:
        anchor_date, rows = _load_scorecard_rows(anchor_str, req_months)
        inventory_trend = get_brand_inventory_trend(
            anchor_date,
            scorecard_rows=rows,
            cache_suffix=f"req:{req_months or 'default'}",
            use_snapshot_query=False,
        )
        return jsonify({"success": True, "trend": inventory_trend})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@inventory_bp.route("/brand-details")
@login_required
def brand_details():
    default_anchor = get_default_anchor_date()
    anchor_str = request.args.get("anchor") or default_anchor.strftime("%Y-%m-%d")
    req_months = (request.args.get("req_months") or "").strip()
    distributor = CITADEL_DISTRIBUTOR
    brand = request.args.get("brand", "").strip()

    try:
        anchor_date = datetime.strptime(anchor_str, "%Y-%m-%d").date()
        anchor_str = anchor_date.strftime("%Y-%m-%d")
        rows = apply_requirement_settings(get_citadel_scorecard_data(anchor_date), CITADEL_DISTRIBUTOR)
        rows = apply_requirement_months(rows, req_months)
        details = get_brand_detail(rows, brand)

        details = sorted(
            details,
            key=lambda r: (
                str(r.get("Style") or ""),
                str(r.get("Colour") or ""),
                str(r.get("SSize") or ""),
            )
        )

        return jsonify({
            "success": True,
            "rows": details
        })

    except Exception as e:
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500


@inventory_bp.route("/kpi-export")
@login_required
def kpi_export():
    anchor_str = request.args.get("anchor")
    req_months = (request.args.get("req_months") or "").strip()
    kpi = (request.args.get("kpi") or "").strip()

    try:
        anchor_date, rows, brand_summary, analysis = _load_scorecard_context(anchor_str, req_months)
        export_rows, export_type = _filter_kpi_export_rows(rows, brand_summary, kpi)
        detail = analysis.get("kpi_details", {}).get(kpi, {})

        if not export_rows:
            flash("No rows found to export for this KPI.", "warning")
            return redirect(url_for("inventory.index", anchor=anchor_date.strftime("%Y-%m-%d"), req_months=req_months))

        summary = analysis.get("summary", {})
        summary_rows = [
            {"Metric": "KPI", "Value": detail.get("title", kpi)},
            {"Metric": "Anchor Date", "Value": anchor_date.strftime("%Y-%m-%d")},
            {"Metric": "Distributor", "Value": CITADEL_DISTRIBUTOR},
            {"Metric": "Requirement Override Months", "Value": req_months or "Seeded / database"},
            {"Metric": "Export Rows", "Value": len(export_rows)},
            {"Metric": "Health Score", "Value": summary.get("health_score", 0)},
            {"Metric": "Average Coverage", "Value": summary.get("avg_coverage", 0)},
            {"Metric": "Available Qty", "Value": summary.get("total_available", 0)},
            {"Metric": "In Transit Qty", "Value": summary.get("total_in_transit", 0)},
            {"Metric": "Required Order Qty", "Value": summary.get("total_req_order", 0)},
            {"Metric": "Critical SKUs", "Value": summary.get("critical_skus", 0)},
            {"Metric": "New SKUs", "Value": summary.get("new_skus", 0)},
            {"Metric": "Below Target SKUs", "Value": summary.get("low_coverage_skus", 0)},
            {"Metric": "Sales 6M", "Value": summary.get("total_sales_6m", 0)},
        ]

        if export_type == "brand":
            detail_rows = [{
                "Brand": row.get("Brand"),
                "Rows": row.get("Rows"),
                "Avg Coverage %": row.get("AvgCoveragePct"),
                "Avg Coverage Incl InTransit %": row.get("AvgCoveragePctInclInTransit"),
                "Critical Count": row.get("CriticalCount"),
                "Low Coverage Count": row.get("LowCoverageCount"),
                "Out Of Stock Count": row.get("OutOfStockCount"),
                "Req Order Count": row.get("ReqOrderCount"),
                "Available Qty": row.get("TotalAvailableQty"),
                "In Transit Qty": row.get("TotalInTransitQty"),
                "Req Order Qty": row.get("TotalReqOrderQty"),
            } for row in export_rows]
        else:
            detail_rows = [{
                "Distributor": row.get("Distributor"),
                "Brand": row.get("Brand"),
                "Style": row.get("Style"),
                "Colour": row.get("Colour"),
                "Size": row.get("SSize"),
                "SKU": row.get("SKU"),
                "Launch Date": row.get("dtColLaunch"),
                "SKU Old Months": row.get("SKUOldMonths"),
                "Stock Qty": row.get("StockQty"),
                "Pipeline Qty": row.get("PipelineQty"),
                "Available Qty": row.get("AvailableQty"),
                "In Transit Qty": row.get("InTransitQty"),
                "Open Orders": row.get("OpenOrders"),
                "Avg Monthly Sales 6M": row.get("AvgMonthlySales6M"),
                "Sales 6M": row.get("Sales6M"),
                "Requirement Months": row.get("StockReqMonths"),
                "Req Stock": row.get("ReqStock"),
                "Coverage %": row.get("CoveragePct"),
                "Coverage Incl InTransit %": row.get("CoveragePct_InclInTransit"),
                "Coverage Months": row.get("CoverageMonths"),
                "Req New Order Qty": row.get("ReqNewOrderQty"),
            } for row in export_rows]

        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            pd.DataFrame(summary_rows).to_excel(writer, index=False, sheet_name="Summary")
            pd.DataFrame(detail_rows).to_excel(writer, index=False, sheet_name="Details")
            _format_inventory_workbook(writer)
        output.seek(0)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_kpi = kpi.replace(" ", "_") or "kpi"
        filename = f"citadel_{safe_kpi}_detail_{timestamp}.xlsx"

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        flash(str(e), "danger")
        return redirect(url_for("inventory.index"))


@inventory_bp.route("/requirements", methods=["GET", "POST"])
@login_required
def requirements():
    distributor = CITADEL_DISTRIBUTOR
    default_anchor = get_default_anchor_date()
    anchor_str = request.args.get("anchor") or default_anchor.strftime("%Y-%m-%d")
    error = None
    rows = []
    brand_options = []

    if request.method == "POST":
        action = request.form.get("action", "save")
        brand = (request.form.get("brand") or "").strip()

        try:
            if action == "delete":
                delete_requirement_setting(distributor, brand)
                flash(f"Requirement rule removed for {brand}.", "success")
            else:
                requirement_months = request.form.get("requirement_months")
                save_requirement_setting(distributor, brand, requirement_months)
                flash(f"Requirement months saved for {brand}.", "success")
        except Exception as e:
            flash(str(e), "danger")

        return redirect(url_for("inventory.requirements", anchor=request.form.get("anchor") or anchor_str))

    try:
        anchor_date = datetime.strptime(anchor_str, "%Y-%m-%d").date()
        anchor_str = anchor_date.strftime("%Y-%m-%d")
        rows = get_citadel_scorecard_data(anchor_date)
        brand_options = build_requirement_brand_options(rows)
    except Exception as e:
        error = str(e)

    settings = get_requirement_settings(distributor)
    settings_by_brand = {
        str(item.get("Brand") or ""): item
        for item in settings
    }

    return render_template(
        "inventory/requirements.html",
        error=error,
        anchor=anchor_str,
        distributor=distributor,
        brand_options=brand_options,
        settings=settings,
        settings_by_brand=settings_by_brand,
        last_sync=datetime.now().strftime("%d-%b-%Y %I:%M %p"),
    )


@inventory_bp.route("/presentation")
@login_required
def presentation():
    error = None
    rows = []
    summary = {}
    mapping_used = {}
    tree_data = []

    distributor = request.args.get("distributor", "").strip()
    brand = request.args.get("brand", "").strip()
    style = request.args.get("style", "").strip()
    risk = request.args.get("risk", "").strip()
    selected_batch = request.args.get("batch_id", "").strip()

    try:
        rows = get_presentation_rows_from_db(selected_batch or None)
        rows = apply_presentation_filters(
            rows,
            distributor=distributor,
            brand=brand,
            style=style,
            risk=risk,
        )
        summary = get_presentation_summary(rows)
        tree_data = build_saved_inventory_brand_tree(rows)
        batches = get_presentation_batches()

        return render_template(
            "inventory/presentation.html",
            rows=rows,
            error=error,
            summary=summary,
            mapping_used=mapping_used,
            distributor=distributor,
            brand=brand,
            style=style,
            risk=risk,
            batches=batches,
            selected_batch=selected_batch,
            tree_data=tree_data,
            last_sync=datetime.now().strftime("%d-%b-%Y %I:%M %p"),
        )

    except Exception as e:
        error = str(e)
        batches = get_presentation_batches()

        return render_template(
            "inventory/presentation.html",
            rows=[],
            error=error,
            summary={},
            mapping_used=mapping_used,
            distributor=distributor,
            brand=brand,
            style=style,
            risk=risk,
            batches=batches,
            selected_batch=selected_batch,
            tree_data=[],
            last_sync=datetime.now().strftime("%d-%b-%Y %I:%M %p"),
        )

@inventory_bp.route("/presentation/export")
@login_required
def presentation_export():
    distributor = request.args.get("distributor", "").strip()
    brand = request.args.get("brand", "").strip()
    style = request.args.get("style", "").strip()
    risk = request.args.get("risk", "").strip()
    selected_batch = request.args.get("batch_id", "").strip()

    try:
        rows = get_presentation_rows_from_db(selected_batch or None)
        rows = apply_presentation_filters(
            rows,
            distributor=distributor,
            brand=brand,
            style=style,
            risk=risk,
        )

        if not rows:
            flash("No rows found to export.", "warning")
            return redirect(url_for("inventory.presentation", batch_id=selected_batch))

        df = pd.DataFrame(rows)

        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Inventory")
        output.seek(0)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"presentation_inventory_{timestamp}.xlsx"

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        flash(str(e), "danger")
        return redirect(url_for("inventory.presentation", batch_id=selected_batch))

@inventory_bp.route("/presentation/compare")
@login_required
def presentation_compare():
    old_batch_id = request.args.get("old_batch_id", "").strip()
    new_batch_id = request.args.get("new_batch_id", "").strip()

    error = None
    comparisons = []
    batches = get_presentation_batches()

    try:
        if old_batch_id and new_batch_id:
            comparisons = compare_presentation_batches(old_batch_id, new_batch_id)
    except Exception as e:
        error = str(e)

    return render_template(
        "inventory/presentation_compare.html",
        error=error,
        comparisons=comparisons,
        batches=batches,
        old_batch_id=old_batch_id,
        new_batch_id=new_batch_id,
        last_sync=datetime.now().strftime("%d-%b-%Y %I:%M %p")
    )

@inventory_bp.route("/presentation/batch/<batch_id>")
@login_required
def presentation_batch(batch_id):
    return redirect(url_for("inventory.presentation", batch_id=batch_id))


@inventory_bp.route("/presentation/delete-batch/<batch_id>", methods=["POST"])
@login_required
def presentation_delete_batch(batch_id):
    try:
        delete_presentation_batch(batch_id)
        flash("Batch deleted successfully.", "success")
    except Exception as e:
        flash(str(e), "danger")

    return redirect(url_for("inventory.presentation_upload"))


@inventory_bp.route("/presentation/upload", methods=["GET", "POST"])
@login_required
def presentation_upload():
    error = None
    uploaded_filename = None

    if request.method == "POST":
        file = request.files.get("inventory_file")
        if not file or file.filename == "":
            error = "Please choose an Excel or CSV file."
        else:
            try:
                upload_dir = os.path.abspath(
                    os.path.join(current_app.root_path, "..", "static", "uploads", "inventory")
                )
                os.makedirs(upload_dir, exist_ok=True)

                filename = secure_filename(file.filename)
                uploaded_filename = filename
                file_path = os.path.join(upload_dir, filename)
                file.save(file_path)

                df = read_uploaded_inventory(file_path)
                standardized_df, _ = standardize_inventory_dataframe(df)
                rows = dataframe_to_records(standardized_df)
                batch_id = save_presentation_rows_to_db(rows, filename)

                flash(f"Inventory uploaded and saved. Batch ID: {batch_id}", "success")
                return redirect(url_for("inventory.presentation", batch_id=batch_id))
            except Exception as e:
                error = str(e)

    batches = get_presentation_batches()
    return render_template(
        "inventory/presentation_upload.html",
        batches=batches,
        error=error,
        uploaded_filename=uploaded_filename,
        last_sync=datetime.now().strftime("%d-%b-%Y %I:%M %p"),
    )
